#! /usr/bin/env python

#
#   To do: put formulae on last line of spreadsheet to total rows, add up delta posn., and exposure
#
# To do: change read_betavals so that it can cope with formulae in the beta column
# (now the .value selector produces a string.
# You need to understand the load_worksheet function better).
#
# Also, TWS handles forex hopelessly: the "Delta Dollars" col. comes out in the foreign currency.
# so a GBP.JPY position is presented as enormous.
# You frigged this by dividing the beta value, but that's wrong.
# maybe the beta value should be zero: are cross rates really solidly correlated with SPY?

import csv
import locale

import openpyxl as opx
import datetime

import sys
import getopt

# import json
# import time
import os


# ALPHAVANTAGE_API_KEY
MY_KEY = '71N6UTNGSMQXQFWU'

vix_level = 0.35 # imp. vol. of SPX. Should be read in

from alpha_vantage.timeseries import TimeSeries

ts = TimeSeries(key=MY_KEY)


locale.setlocale( locale.LC_ALL, 'en_US.UTF-8' )  # so you can interpret e.g. "1,000.0" as a number

# import string


# cols in export from IB
ib_export = {'ticker':1, 'sec_type':2, 'exchange':3, 'datestr':4, 'strike':5, 'putcall':6, 'size':7}


# will hold beta for stock tickers as dict

verbose = False

exclude_sec_types = set() # comma-sep list of sec types to be excluded

def quiet_print(*args, force=False):
    if verbose or force:
        print(args)

tickers = dict()

def read_positions_old():
    positions = []
    with open("example.csv", newline='') as csvfile:
        position_reader = csv.reader(csvfile, dialect='excel', quoting=csv.QUOTE_NUMERIC)
        for row in position_reader:
            positions.append(row)
    return positions

def read_positions(f):
    positions = [] # each entry will be of type dict
    with open(f, newline = '') as csvfile:
        header = [h.strip() for h in csvfile.readline().split(',')]
        quiet_print(header)
        reader = csv.DictReader(csvfile, fieldnames=header)
        for row in reader:
            positions.append(row)
    return positions

def create_betas():
    with open("betas.csv", 'w', newline = '') as betacsvfile:
        betawriter = csv.writer(betacsvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        for i in sorted(tickers):
            betawriter.writerow([i]+[1,2,3,4])

def read_betas(): # read csv file with betas in
    result = dict()
    with open("betas.csv", newline = '') as bcsv:
        breader = csv.reader(bcsv)
        for row in breader:
            result[row[0]] = row[1:]
    return result

def display_betas(betas):
    for i in betas:
        quiet_print(i, betas[i])

def display_positions(p):
    for i in p:
        quiet_print (i)


def read_known_betas(f):  # read additional betas from excel worksheet
    wb = opx.load_workbook(f)
    ws = wb.active # there will be just one
    betavals = dict()
    for row in ws.rows:
        if (row[0].value == "Ticker"): # i.e. 1st row
            continue
        betavals[row[0].value] = [row[1].value, row[2].value] # i.e. beta and sectype
    return betavals


def write_row(ws,row_number, col_start, lst):
    c = col_start
    for i in lst:
        ws.cell(row_number, c).value = i
        c += 1

def atof(cell):
    value = 0.
    try:
        value = locale.atof(cell)
    except:
        quiet_print("failed to convert {0:}".format(cell)) 
    return value

##def get_price_from_alpha_vantage(underlying):
##    try:
##        quotevals = ts.get_quote_endpoint(symbol=underlying)
##        time.sleep(15) # free lib wants fewer than five calls per minute
##        # quote_dict = json.loads(quotevals)
##        print("Symbol {0:} price {1:}".format(underlying, quotevals[0]['05. price']))
##    except ValueError:
##        print("can't get price for {0:}".format(underlying))
##    return quotevals
##        

def usage():
    print("calc_deltas --betafile=<betas.xlsx> --positionfile=<positions.csv>")

def main(argv):
    global verbose # if you change a variable in a function, it's assumed to be local unless explicitly declared
    global vix_level
    global exclude_sec_types
    beta_workbook =     "betavals.xlsx"
    positions_file =    'positions.csv'
    try:
        opts, args = getopt.getopt(sys.argv[1:], "hvb:p:",["verbose", "betafile=",
                                                           "positionfile=", "vix=", "exclude="])
    except getopt.GetoptError as err:
        print(err)
        usage()
        sys.exit(2)
    output = None
    for o, a in opts:
        if o in ("-v", "--verbose"):
            verbose = True
        elif o in ("-h", "--help"):
            usage()
            sys.exit()
        elif o in ("-b", "--betafile"):
            beta_workbook = a
        elif o in ("-p", "--positionfile"):
            positions_file = a
        elif o in ("--vix"):
            vix_level = float(a)/100.0
            quiet_print("using vix level of {0:}".format(vix_level))
        elif o in ("--exclude"):
            exclude_sec_types = a.split(",")
            quiet_print("excluding: {0:}".format(exclude_sec_types))
        else:
            assert False, "unhandled option"
    # betas = read_betas(beta_workbook)
    # can display_betas to check
    positions = read_positions(positions_file)

    known_betas = read_known_betas(beta_workbook)
    quiet_print("known betas: {}".format(known_betas), force=False)
    display_positions(positions)
    aggregate_delta = 0
    wb = opx.Workbook() # workbook
    ws = wb.create_sheet('Delta Calc', 0)
    excel_row = 1
    write_row(ws, excel_row, 1, ["Underlying", "Instrument", "Beta", "Delta",
                                 "Position Delta", "Exposure", "Volatility"])
    excel_row += 1
    # you need to move this big loop into a subroutine which returns a dict of all the values you need.
    for i in positions:
        fi = i['Financial Instrument']
        quiet_print("{0:}".format(fi))
        underlying = i['Underlying']
        beta = i['Beta']
        delta_str = i['Delta Dollars']
        delta = atof(delta_str)
        if delta == 0.:
            print("zero delta for {0:} -- is that right? Will use mkt value.".format(fi))
            delta = atof(i['Market Value'])
        try:
            delta = locale.atof(delta_str)
        except ValueError:
            print("{2:}: couldn't convert {0:} to float, using mkt val: {1:}".format(delta_str, delta, fi))
            
        quiet_print("delta={0:} of type {1:}".format(delta, type(delta)))
        volstr = i['Closing Impl. Vol. %']
        quiet_print("vol is {0:}".format(volstr))
        # n.b. b & d are strings
        quiet_print("{0:}, beta={1:}, delta={2:}".format(i['Financial Instrument'],beta, delta))
        betav = atof(beta) # local atof uses locale.atof
        if betav == 0.:
            if underlying in known_betas:
                if known_betas[underlying][1] in exclude_sec_types:
                    betav = 0.
                else:
                    betav = known_betas[underlying][0]
            else:
                print("can't read beta for: {0:} -- assuming 1".format(i))
                betav = 1.0
        
        if volstr == "" or volstr == "NoMD":
            volstr = i['Hist. Vol. %']
        quiet_print("hist vol is {0:}".format(volstr))
        try:
            vol = float(volstr.strip("%"))/100.
        except ValueError:
            quiet_print("no hist vol for {0:}".format(i['Underlying']))
            vol = vix_level * betav # not 100% sure this makes sense. 
            
        try:
            delta_pos = betav * delta
        except ValueError:
            print("something wrong with betav={0:} or delta={1:}".format(betav, delta))
            

        quiet_print("{0:}, {1:.0f}".format(fi, delta_pos))
        write_row(ws, excel_row, 1, [i["Underlying"], fi, betav,
                                         i["Delta"], delta, delta_pos, vol])
        excel_row += 1

        aggregate_delta += delta_pos
        # end of loop
        
    DV_root = "Delta_Values_"
    gen_filename = DV_root+datetime.datetime.now().strftime("%H%M%S")+".xlsx"
    wb.save(gen_filename)
    link_name = DV_root+'latest.xlsx'
    try:
        os.remove(link_name)
    except OSError:
        print("error deleting {0:} (is it open in Excel?)".format(link_name))
        return False
    os.link(gen_filename, DV_root+'latest.xlsx')
    print("Deltas saved to {0:}".format(gen_filename))
    print("Deltas hard linked to {0:}".format(link_name))
    print("Overall exposure: ${0:n}".format(aggregate_delta))
    return True

if __name__ == "__main__":
    if main(sys.argv[1:]):
        key_word = "without"
    else:
        key_word = "with"
    print("calculation finished {0:} errors".format(key_word))

